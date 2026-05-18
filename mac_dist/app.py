from __future__ import annotations

# 技術経営専門職学位課程 修了要件チェッカー
#
# このファイルは、1ファイルで完結するローカルWebアプリです。
# 処理の流れは大きく次の5段階です。
#
# 1. ブラウザでExcelファイルを選択する
# 2. HTTP POSTでPythonサーバーへExcelを送る
# 3. openpyxlで各シートの科目行を Course オブジェクトへ変換する
# 4. 科目コード・科目名から各科目を修了要件のカテゴリへ自動配分する
# 5. シートごとに単位要件を集計し、JSONでブラウザへ返して表示する
#
# 注意:
# - H列の手入力分類は利用しない前提です。現在の画面は常に自動配分を使います。
# - プロジェクトレポート審査と最終試験はExcelの単位データだけでは判定できません。
# - 科目属性の外部マスタは持っておらず、このファイル内の規則で判定します。

import html
import io
import json
import re
import socket
import threading
import unicodedata
import webbrowser
from dataclasses import dataclass, field
from email import policy
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

import openpyxl
from pypdf import PdfReader


# 自動配分後に course.category へ入ることを許可するカテゴリ一覧です。
# ここにない値は集計対象外にします。
# 例:
# - "B" はB群の必要4単位を満たすために使うカテゴリ
# - "専門" はA/B/C/D/Eの個別要件を満たした後の標準学修課程内の余剰単位などを入れるカテゴリ
ALLOWED_CATEGORIES = {
    "文系教養400",
    "文系教養500",
    "アントレGA0M",
    "アントレGA1M",
    "A400",
    "B",
    "C",
    "D/E",
    "専門",
    "講究",
}

DISPLAY_CATEGORY_ORDER = {
    "文系教養400": 0,
    "文系教養500": 1,
    "アントレGA0M": 2,
    "アントレGA1M": 3,
    "A400": 4,
    "B": 5,
    "C": 6,
    "D/E": 7,
    "講究": 8,
    "専門": 9,
}

DISPLAY_CATEGORY_LABELS = {
    "文系教養400": "文系教養400",
    "文系教養500": "文系教養500",
    "アントレGA0M": "GA0M",
    "アントレGA1M": "GA1M",
    "A400": "A群400番台",
    "B": "B群",
    "C": "C群",
    "D/E": "D群またはE群",
    "講究": "講究科目",
    "専門": "それ以外（専門）",
}

# 修了判定で画面に表示する単位要件です。
# 各タプルは (表示名, 集計キー, 必要単位) です。
#
# 集計キーがカテゴリ名の場合:
#   totals[カテゴリ名] をそのまま見ます。
#
# 特殊キー:
# - "total": すべての科目の総単位
# - "standard_specialty": A400 + B + C + D/E + 専門 の合計
REQUIREMENTS = [
    ("総単位", "total", 40),
    ("文系教養科目 400番台", "文系教養400", 2),
    ("文系教養科目 500番台", "文系教養500", 1),
    ("アントレプレナーシップ科目 GA0M", "アントレGA0M", 1),
    ("アントレプレナーシップ科目 GA1M", "アントレGA1M", 1),
    ("講究科目", "講究", 8),
    ("A群 400番台", "A400", 4),
    ("B群", "B", 4),
    ("C群", "C", 4),
    ("D群またはE群", "D/E", 2),
    ("専門科目群（標準学修課程）", "standard_specialty", 25),
]


@dataclass
class Course:
    """Excelの1行から読み取った1科目分のデータ。

    Excelには多くの見出し・注記・集計表も混ざるため、read_courses() では
    科目コードが TIM. または LAH. で始まる行だけを Course として扱います。

    category は自動配分後に入る「この科目をどの要件へ充当したか」の値です。
    source は画面の「根拠」列に出す説明です。
    """

    sheet: str
    row: int
    raw_type: str
    code: str
    name: str
    credit_text: str
    term: str
    assigned: str | None
    credits: float
    category: str | None = None
    source: str = "未分類"
    notes: list[str] = field(default_factory=list)

    @property
    def level(self) -> int | None:
        # 科目コードから 400/500 などの番台を取り出します。
        # TIM.A401 -> 401, LAH.S433 -> 433 のような値になります。
        match = re.search(r"\.(?:[A-Z])?(\d{3})", self.code)
        if not match:
            match = re.search(r"(\d{3})", self.code)
        return int(match.group(1)) if match else None

    @property
    def group(self) -> str | None:
        # TIM.A401 の A、TIM.B410 の B のように、TIM系科目の群を取り出します。
        # LAH科目にはA/B/C/D/E群がないため None になります。
        match = re.search(r"TIM\.([A-Z])", self.code)
        return match.group(1) if match else None

    @property
    def is_tim(self) -> bool:
        return self.code.startswith("TIM.")

    @property
    def is_lah(self) -> bool:
        return self.code.startswith("LAH.")

    @property
    def has_ga0(self) -> bool:
        # アントレプレナーシップ要件は科目名中の GA 表記から判定します。
        return "GA0M" in self.name

    @property
    def has_ga1(self) -> bool:
        # GA1Mも同様に科目名中の表記を見ます。
        return "GA1M" in self.name

    @property
    def liberal_compatible(self) -> bool:
        # 文系教養科目として扱える科目かどうかを判定します。
        # LAH科目は文系教養科目です。
        # TIM科目でも「文系教養対応科目」と科目名にあるものは候補に含めます。
        return self.is_lah or "文系教養対応科目" in self.name

    @property
    def standard_eligible(self) -> bool:
        # 標準学修課程の専門科目群に含める候補です。
        # TIMのA/B/C/D/E群を対象にしています。
        return self.group in {"A", "B", "C", "D", "E"}


def normalize_text(value) -> str:
    """Excelセル値を空文字または前後空白を除いた文字列へそろえる。"""
    if value is None:
        return ""
    return str(value).strip()


def parse_credit(value, code: str) -> float:
    """Excelの単位欄から単位数を計算する。

    学修案内・サンプルExcelでは、単位欄が "1-0-0" や "0.5-0.5-0" のように
    講義・演習・実験等の内訳で入っているため、それらの合計を単位数にします。

    講究科目はサンプルによって日付形式に化けて読まれることがあったため、
    対象コードは明示的に2単位として扱います。
    それ以外で解釈できない場合は保守的に1単位とします。
    """
    text = normalize_text(value)
    if re.fullmatch(r"\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?){2}", text):
        return sum(float(part) for part in text.split("-"))
    if code in {"TIM.Z491", "TIM.Z492", "TIM.Z591", "TIM.Z592"}:
        return 2.0
    return 1.0


def compact_number(value: float) -> int | float:
    """画面表示用に、整数相当の 4.0 を 4 として返す。"""
    return int(value) if abs(value - round(value)) < 0.0001 else round(value, 2)


def read_courses(file_bytes: bytes, filename: str = "") -> list[Course]:
    """アップロードファイルの形式を見て、Excel/PDFの読み取り処理へ振り分ける。"""
    lower_name = filename.lower()
    if file_bytes.startswith(b"%PDF") or lower_name.endswith(".pdf"):
        return read_courses_from_pdf(file_bytes)
    if lower_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")) or file_bytes.startswith(b"PK"):
        return read_courses_from_excel(file_bytes)
    raise ValueError("ExcelまたはPDFファイルを選択してください。")


def read_courses_from_excel(file_bytes: bytes) -> list[Course]:
    """アップロードされたExcelから科目行だけを読み取る。

    想定している主な列:
    - A列: 科目区分の表示
    - B列: 科目コード
    - C列: 科目名
    - E列: 単位数または単位内訳
    - G列: 開講学期
    - H列: 手入力分類

    現在のUIはH列を使わない「自動配分」固定ですが、過去の検証用に
    assigned へは値を保持しています。
    """
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    courses: list[Course] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in range(1, ws.max_row + 1):
            code = normalize_text(ws.cell(row, 2).value)
            # 見出し・集計表・空行を除外し、科目コードらしい行だけを採用します。
            if not re.match(r"^(TIM|LAH)\.", code):
                continue
            # 一部のExcelでは "TIM.A.401" のように余分なドットが入る可能性があるため、
            # 後続の正規表現で扱いやすい "TIM.A401" 形式へ寄せます。
            code = code.split()[0].replace(".A.", ".").replace(".B.", ".").replace(".C.", ".")
            credit_cell = ws.cell(row, 5).value
            assigned = normalize_text(ws.cell(row, 8).value) or None
            name = merge_continuation_notes(ws, row, normalize_text(ws.cell(row, 3).value))
            courses.append(
                Course(
                    sheet=sheet_name,
                    row=row,
                    raw_type=normalize_text(ws.cell(row, 1).value),
                    code=code,
                    name=name,
                    credit_text=normalize_text(credit_cell),
                    term=normalize_text(ws.cell(row, 7).value),
                    assigned=assigned if assigned in ALLOWED_CATEGORIES else None,
                    credits=parse_credit(credit_cell, code),
                )
            )
    return courses


def merge_continuation_notes(ws, row: int, name: str) -> str:
    """科目名の次行に分かれた属性注記を科目名へ結合する。

    新井単位.xlsx のようなデータでは、1科目が次のように2行で表現されます。

    - 1行目: TIM.A405 / 数理情報分析基礎 I / 0.5-0.5-0
    - 2行目: 科目コードなし / (GA：GA0M)

    既存ロジックは「科目コードがある行だけ」を科目として読むため、
    2行目をそのまま無視すると GA0M や 文系教養対応科目 の属性を失います。
    そこで、直後に科目コードがなく、C列に括弧書きの属性注記がある場合だけ、
    その文字列を科目名へ足してから分類します。
    """
    notes: list[str] = []
    next_row = row + 1
    while next_row <= ws.max_row:
        next_code = normalize_text(ws.cell(next_row, 2).value)
        next_name = normalize_text(ws.cell(next_row, 3).value)
        if next_code or not next_name:
            break
        if not (next_name.startswith("(") or next_name.startswith("（")):
            break
        notes.append(next_name)
        next_row += 1
    return " ".join([name, *notes]).strip()


COURSE_CODE_RE = re.compile(r"\b(?:TIM|LAH)\.[A-Z]?\d{3}\b")
CREDIT_PATTERN_RE = re.compile(r"\d+(?:\.\d+)?-\d+(?:\.\d+)?-\d+(?:\.\d+)?")
PDF_META_RE = re.compile(r"学籍番号\s*(\S+).*?氏名\s*(.+)")
PDF_SKIP_PHRASES = (
    "推奨 科目コー",
    "ド",
    "授業科目名",
    "専門科目",
    "文系教養科目",
    "講究科目",
    "修得済単位",
    "修士課程",
    "2026年",
)


def read_courses_from_pdf(file_bytes: bytes) -> list[Course]:
    """教務Webシステムの成績一覧PDFから科目を抽出する。

    PDF抽出テキストは行分割が不安定なため、科目コードを起点に1科目分の
    ブロックを集め、ブロック内の単位内訳や属性注記を拾ってCourse化します。
    """
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
    except Exception as exc:
        raise ValueError("PDFを読み取れませんでした。教務Webシステムの成績一覧PDFか確認してください。") from exc
    pages = []
    for page_number, page in enumerate(reader.pages, start=1):
        text = unicodedata.normalize("NFKC", page.extract_text() or "")
        pages.append((page_number, [line.strip() for line in text.splitlines() if line.strip()]))

    if not any(lines for _, lines in pages):
        raise ValueError("PDFから文字を抽出できませんでした。Excel形式の入力をお試しください。")

    sheet_name = pdf_sheet_name(pages)
    blocks = collect_pdf_course_blocks(pages)
    courses = [course_from_pdf_block(sheet_name, block) for block in blocks]
    courses = [course for course in courses if course is not None]
    if not courses:
        raise ValueError("PDFから科目コードを抽出できませんでした。教務Webシステムの成績一覧PDFか確認してください。")
    return courses


def pdf_sheet_name(pages) -> str:
    """PDF内の学籍番号・氏名が抽出できれば、画面の見出しに使う。"""
    first_text = " ".join(pages[0][1]) if pages else ""
    match = PDF_META_RE.search(first_text)
    if not match:
        return "PDF"
    student_id = match.group(1).strip()
    name = re.split(r"\s+修得済単位|\s+合計", match.group(2).strip())[0].strip()
    return f"PDF {student_id} {name}".strip()


def collect_pdf_course_blocks(pages) -> list[dict]:
    """科目コード行から次の科目コード行までを1ブロックとして集める。"""
    blocks: list[dict] = []
    current: dict | None = None
    for page_number, lines in pages:
        for line in lines:
            if should_skip_pdf_line(line):
                continue
            if COURSE_CODE_RE.search(line):
                if current:
                    blocks.append(current)
                current = {"page": page_number, "lines": [line]}
            elif current:
                current["lines"].append(line)
    if current:
        blocks.append(current)
    return blocks


def should_skip_pdf_line(line: str) -> bool:
    """ヘッダーやページ下部メッセージなど、科目ブロックに混ぜない行を除外する。"""
    if line in {"A", "B", "C", "D", "E", "L", "R", "(○)", "(◎)", "(選", "択)", "-"}:
        return True
    return any(phrase in line for phrase in PDF_SKIP_PHRASES)


def course_from_pdf_block(sheet_name: str, block: dict) -> Course | None:
    """PDFの1科目ブロックをCourseへ変換する。"""
    lines = block["lines"]
    text = " ".join(lines)
    code_match = COURSE_CODE_RE.search(text)
    if not code_match:
        return None

    code = normalize_course_code(code_match.group(0))
    after_code = text[code_match.end():].strip()
    credit_match = CREDIT_PATTERN_RE.search(after_code)
    credit_text = credit_match.group(0) if credit_match else ""
    name_area = after_code[: credit_match.start()].strip() if credit_match else after_code
    name = clean_pdf_course_name(name_area)
    term = extract_pdf_term(text)
    return Course(
        sheet=sheet_name,
        row=block["page"],
        raw_type="PDF",
        code=code,
        name=name,
        credit_text=credit_text,
        term=term,
        assigned=None,
        credits=parse_credit(credit_text, code),
    )


def normalize_course_code(code: str) -> str:
    """科目コード表記を既存ロジックが扱いやすい形へ寄せる。"""
    return code.split()[0].replace(".A.", ".").replace(".B.", ".").replace(".C.", ".")


def clean_pdf_course_name(value: str) -> str:
    """PDF抽出テキストから教員名などを完全には消さず、属性注記を残した科目名に整える。"""
    value = re.sub(r"\s+", " ", value).strip()
    return value


def extract_pdf_term(text: str) -> str:
    """PDFブロックから修得時期を取り出す。"""
    compact = text.replace(" ", "")
    match = re.search(r"20\d{2}(?:前|後)学期", compact)
    return match.group(0) if match else ""


def assign_from_sheet(courses: list[Course]) -> None:
    """H列の分類を優先してカテゴリを付ける旧モード。

    現在の画面ではH列は空という前提のため使っていません。
    ただし、検証や将来の切替に備えて関数は残しています。
    """
    for course in courses:
        if course.assigned:
            course.category = course.assigned
            course.source = "Excel H列"
        else:
            course.category = infer_single_category(course)
            course.source = "自動推定" if course.category else "未分類"


def infer_single_category(course: Course) -> str | None:
    """1科目だけを見て、もっとも自然なカテゴリを推定する。

    この関数は「その科目が何者か」を単純判定する補助関数です。
    卒業要件全体を満たすための最適な割り当ては auto_allocate() が行います。

    例:
    - LAH.S433 -> 文系教養400
    - TIM.B410 -> B
    - 科目名に GA0M -> アントレGA0M
    - TIM.Z491 -> 講究
    """
    if course.code in {"TIM.Z491", "TIM.Z492", "TIM.Z591", "TIM.Z592"}:
        return "講究"
    if course.is_lah and course.level:
        return "文系教養500" if course.level >= 500 else "文系教養400"
    if course.has_ga0:
        return "アントレGA0M"
    if course.has_ga1:
        return "アントレGA1M"
    if course.group == "A" and course.level and course.level < 500:
        return "A400"
    if course.group == "B":
        return "B"
    if course.group == "C":
        return "C"
    if course.group in {"D", "E"}:
        return "D/E"
    if course.standard_eligible:
        return "専門"
    return None


def auto_allocate(courses: list[Course]) -> None:
    """シート内の科目を、修了要件を満たしやすい順に自動配分する。

    重要な考え方:
    - 1科目は1つのカテゴリにだけ充当します。
    - 個別要件は必要単位を満たすところまで充当します。
      たとえばB群が6単位あっても、B欄には必要単位の4単位分だけ入れます。
    - 個別要件を満たした後に残ったA/B/C/D/E群の科目は "専門" に入れます。
      これにより「専門科目群（標準学修課程）25単位」の合計へ回ります。
    - H列の入力は見ません。ユーザー要件に合わせ、科目コード・科目名だけで決めます。
    """
    for course in courses:
        # 再判定時に前回のカテゴリが残らないよう、最初に全科目を未割当へ戻します。
        course.category = None
        course.source = "自動配分"

    def available(predicate):
        # まだカテゴリが決まっていない科目だけを候補にします。
        # これにより1科目が複数要件へ二重計上されることを防ぎます。
        return [c for c in courses if c.category is None and predicate(c)]

    def take(category: str, needed: float, candidates: list[Course]) -> None:
        # category の必要単位を満たすまで、候補科目を順に割り当てます。
        # 必要単位に到達した時点で止めるため、個別要件の表示は上限カットされます。
        got = sum(c.credits for c in courses if c.category == category)
        for course in candidates:
            if got >= needed:
                break
            course.category = category
            got += course.credits

    # 講究は専門科目群の内訳とは別に8単位要件があるため、先に確保します。
    for c in available(lambda x: x.code in {"TIM.Z491", "TIM.Z492", "TIM.Z591", "TIM.Z592"}):
        c.category = "講究"

    # 教養科目群の個別要件を先に満たします。
    # 文系教養はLAH科目だけでなく、科目名に「文系教養対応科目」とあるTIM科目も候補です。
    take("文系教養400", 2, available(lambda x: x.liberal_compatible and (x.level or 0) < 500))
    take("文系教養500", 1, available(lambda x: x.liberal_compatible and (x.level or 0) >= 500))
    take("アントレGA0M", 1, available(lambda x: x.has_ga0))
    take("アントレGA1M", 1, available(lambda x: x.has_ga1))

    # 標準学修課程の専門科目群に関する個別要件です。
    # A群は400番台指定があるため A400 として分けています。
    take("A400", 4, available(lambda x: x.group == "A" and (x.level or 0) < 500))
    take("B", 4, available(lambda x: x.group == "B"))
    take("C", 4, available(lambda x: x.group == "C"))
    take("D/E", 2, available(lambda x: x.group in {"D", "E"}))

    # 上の個別要件に使わなかったA/B/C/D/E群の科目を、
    # 標準学修課程の専門科目群25単位にカウントするため "専門" へ入れます。
    for course in available(lambda x: x.standard_eligible):
        course.category = "専門"

    # ここまで未割当のLAH科目は、個別要件の上限を超えた文系教養として扱います。
    # 総単位・分類済み単位には反映されますが、個別要件の不足計算はすでに満たされています。
    for course in available(lambda x: x.is_lah and x.level and x.level < 500):
        course.category = "文系教養400"
    for course in available(lambda x: x.is_lah and x.level and x.level >= 500):
        course.category = "文系教養500"


def evaluate_sheet(courses: list[Course], mode: str) -> dict:
    """1シート分の科目を集計し、画面表示用の判定結果を作る。

    サンプルExcelでは学生ごとにシートが分かれているため、
    evaluate() でシート別に分けた後、この関数を呼びます。
    """
    if mode == "auto":
        auto_allocate(courses)
    else:
        assign_from_sheet(courses)

    totals = {category: 0.0 for category in ALLOWED_CATEGORIES}
    for course in courses:
        if course.category in totals:
            totals[course.category] += course.credits

    # total_credits はカテゴリ配分に関係なく、読み取った全科目の単位合計です。
    total_credits = sum(course.credits for course in courses)

    # 標準学修課程の専門科目群25単位は、A400/B/C/D-Eの個別充当分と、
    # 余剰専門科目を入れた "専門" の合計で判定します。
    standard_specialty = sum(totals[key] for key in ["A400", "B", "C", "D/E", "専門"])

    # 上部サマリー用の合計です。卒業判定の本体は REQUIREMENTS の行で行います。
    liberal_total = totals["文系教養400"] + totals["文系教養500"]
    entrepreneurship_total = totals["アントレGA0M"] + totals["アントレGA1M"]

    rows = []
    for label, key, required in REQUIREMENTS:
        # REQUIREMENTS の key に応じて、見るべき実績値を切り替えます。
        actual = total_credits if key == "total" else standard_specialty if key == "standard_specialty" else totals[key]
        rows.append(
            {
                "label": label,
                "actual": compact_number(actual),
                "required": required,
                "ok": actual + 1e-9 >= required,
                "shortage": compact_number(max(0, required - actual)),
            }
        )

    overall_ok = all(row["ok"] for row in rows)
    # 戻り値はそのままJSON化してブラウザに渡すため、辞書と配列だけで構成します。
    return {
        "sheetName": courses[0].sheet if courses else "データなし",
        "overallOk": overall_ok,
        "overallShortage": compact_number(max(0, 40 - total_credits)),
        "summary": {
            "total": compact_number(total_credits),
            "liberal": compact_number(liberal_total),
            "entrepreneurship": compact_number(entrepreneurship_total),
            "standardSpecialty": compact_number(standard_specialty),
            "classified": compact_number(sum(c.credits for c in courses if c.category)),
            "courseCount": len(courses),
        },
        "requirements": rows,
        "courses": [
            {
                "sheet": c.sheet,
                "row": c.row,
                "code": c.code,
                "name": c.name,
                "credits": compact_number(c.credits),
                "term": c.term,
                "category": display_category_label(c.category),
                "source": c.source,
            }
            for c in sorted(courses, key=course_display_sort_key)
        ],
    }


def course_display_sort_key(course: Course) -> tuple[int, str, int, str]:
    """画面の科目一覧を、利用者が確認しやすい充当先順に並べる。"""
    return (
        DISPLAY_CATEGORY_ORDER.get(course.category or "", 99),
        course.sheet,
        course.row,
        course.code,
    )


def display_category_label(category: str | None) -> str:
    """内部カテゴリ名を画面表示用の名称へ変換する。"""
    if not category:
        return "未分類"
    return DISPLAY_CATEGORY_LABELS.get(category, category)


def evaluate(courses: list[Course], mode: str) -> dict:
    """全シートの科目をシート名で分け、シートごとに判定する。

    1つのExcelに複数人分のシートがある場合、合算すると判定が甘くなるため、
    必ずシート単位で独立して判定します。
    """
    grouped: dict[str, list[Course]] = {}
    for course in courses:
        grouped.setdefault(course.sheet, []).append(course)
    sheets = [evaluate_sheet(sheet_courses, mode) for sheet_courses in grouped.values()]
    return {
        "sheets": sheets,
        "notes": [
            "プロジェクトレポートの審査および最終試験の合否は、修得単位データからは判定できません。",
            "H列などの手入力分類は使わず、科目コード・科目名から要件を満たしやすいように自動配分します。",
        ],
    }


# ブラウザへ返すHTML/CSS/JavaScriptです。
# 配布を簡単にするため、テンプレートファイルを別に置かず、
# Python文字列としてこのファイル内に埋め込んでいます。
#
# フロントエンド側の役割:
# - Excelファイル選択UIを表示する
# - /check にファイルをPOSTする
# - Python側から返ってきたJSONを表・サマリーとして描画する
HTML_PAGE = """<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>技術経営専門職学位課程 修了要件チェッカー</title>
  <style>
    :root { color-scheme: light; --ink:#172026; --muted:#5f6f7a; --line:#d8e1e8; --ok:#0f7b53; --ng:#bd3b34; --bg:#f6f8f9; --panel:#fff; --accent:#176b87; }
    * { box-sizing: border-box; }
    body { margin:0; font-family: "Segoe UI", system-ui, sans-serif; background:var(--bg); color:var(--ink); }
    header { background:linear-gradient(135deg,#164f63,#2f7d6a); color:#fff; padding:28px max(24px,6vw); }
    h1 { margin:0; font-size:clamp(24px,3vw,38px); letter-spacing:0; }
    header p { margin:10px 0 0; max-width:880px; color:#e8f4f1; line-height:1.65; }
    main { padding:28px max(20px,6vw) 48px; }
    .toolbar { display:grid; grid-template-columns: 1fr auto; gap:16px; align-items:end; border-bottom:1px solid var(--line); padding-bottom:20px; }
    label { display:block; font-weight:700; margin-bottom:8px; }
    input[type=file] { width:100%; padding:12px; background:#fff; border:1px solid var(--line); border-radius:8px; }
    .primary { border:0; border-radius:6px; padding:11px 14px; font-weight:700; cursor:pointer; }
    .primary { background:var(--accent); color:#fff; min-width:120px; }
    .actions { display:flex; gap:10px; align-items:center; }
    .hero { display:grid; grid-template-columns:minmax(260px,1.2fr) repeat(4,minmax(120px,1fr)); gap:12px; margin:22px 0; }
    .tile { background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:16px; min-height:94px; }
    .tile strong { display:block; font-size:28px; margin-top:6px; }
    .status { border-left:8px solid var(--muted); }
    .status.ok { border-left-color:var(--ok); }
    .status.ng { border-left-color:var(--ng); }
    .status .result { font-size:24px; font-weight:800; }
    .muted { color:var(--muted); font-size:13px; line-height:1.55; }
    table { width:100%; border-collapse:collapse; background:#fff; border:1px solid var(--line); border-radius:8px; overflow:hidden; }
    th, td { padding:11px 12px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; }
    th { background:#eef4f5; font-size:13px; color:#31434d; }
    tr:last-child td { border-bottom:0; }
    .badge { display:inline-block; min-width:42px; text-align:center; border-radius:999px; padding:4px 10px; font-weight:800; font-size:12px; }
    .badge.ok { background:#dff3e9; color:var(--ok); }
    .badge.ng { background:#fae2df; color:var(--ng); }
    .grid { display:grid; grid-template-columns:1fr; gap:20px; }
    .panel-title { margin:26px 0 10px; font-size:18px; }
    .error { margin-top:16px; color:var(--ng); font-weight:700; }
    .notes { margin-top:16px; padding:14px 18px; border:1px solid var(--line); background:#fff; border-radius:8px; }
    @media (max-width: 900px) { .toolbar, .hero { grid-template-columns:1fr; } .actions { flex-wrap:wrap; } table { font-size:14px; } }
  </style>
</head>
<body>
  <header>
    <h1>技術経営専門職学位課程 修了要件チェッカー</h1>
    <p>2026年度学修案内の修了要件に基づき、Excelの修得科目データから単位充足状況を判定します。</p>
  </header>
  <main>
    <section class="toolbar">
      <div>
        <label for="file">チェック用Excel/PDFファイル</label>
        <input id="file" type="file" accept=".xlsx,.xlsm,.pdf">
      </div>
      <div class="actions">
        <button id="run" class="primary" type="button">判定</button>
      </div>
    </section>
    <div id="error" class="error"></div>
    <section id="output"></section>
  </main>
  <script>
    document.getElementById("run").addEventListener("click", async () => {
      const file = document.getElementById("file").files[0];
      const error = document.getElementById("error");
      error.textContent = "";
      if (!file) { error.textContent = "ExcelまたはPDFファイルを選択してください。"; return; }
      const form = new FormData();
      form.append("file", file);
      form.append("mode", "auto");
      const response = await fetch("/check", { method: "POST", body: form });
      const data = await response.json();
      if (!response.ok) { error.textContent = data.error || "判定できませんでした。"; return; }
      render(data);
    });
    function mark(ok) { return `<span class="badge ${ok ? "ok" : "ng"}">${ok ? "OK" : "不足"}</span>`; }
    function esc(value) { return String(value ?? "").replace(/[&<>"']/g, (s) => ({ "&":"&amp;", "<":"&lt;", ">":"&gt;", '"':"&quot;", "'":"&#39;" }[s])); }
    function render(data) {
      const sheets = data.sheets.map((sheet) => {
        const statusClass = sheet.overallOk ? "ok" : "ng";
        const message = sheet.overallOk ? "修了要件を満たしています" : `修了要件まであと ${sheet.overallShortage} 単位`;
        const reqRows = sheet.requirements.map((r) => `<tr><td>${esc(r.label)}</td><td>${r.actual} / ${r.required}</td><td>${r.shortage}</td><td>${mark(r.ok)}</td></tr>`).join("");
        const courseRows = sheet.courses.map((c) => `<tr><td>${esc(c.code)}</td><td>${esc(c.name)}</td><td>${c.credits}</td><td>${esc(c.category)}</td><td>${esc(c.source)}</td><td>${esc(c.sheet)}:${c.row}</td></tr>`).join("");
        return `
        <h2 class="panel-title">${esc(sheet.sheetName)}</h2>
        <section class="hero">
          <div class="tile status ${statusClass}"><div class="muted">総合判定</div><div class="result">${esc(message)}</div></div>
          <div class="tile"><div class="muted">総単位</div><strong>${sheet.summary.total}</strong></div>
          <div class="tile"><div class="muted">教養科目群</div><strong>${sheet.summary.liberal + sheet.summary.entrepreneurship}</strong></div>
          <div class="tile"><div class="muted">専門科目群</div><strong>${sheet.summary.standardSpecialty}</strong></div>
          <div class="tile"><div class="muted">科目数</div><strong>${sheet.summary.courseCount}</strong></div>
        </section>
        <h3 class="panel-title">単位要件ごとの判定</h3>
        <table><thead><tr><th>要件</th><th>修得 / 必要</th><th>不足</th><th>判定</th></tr></thead><tbody>${reqRows}</tbody></table>
        <h3 class="panel-title">科目の充当先</h3>
        <table><thead><tr><th>科目コード</th><th>科目名</th><th>単位</th><th>充当先</th><th>根拠</th><th>位置</th></tr></thead><tbody>${courseRows}</tbody></table>
      `}).join("");
      document.getElementById("output").innerHTML = sheets + `<div class="notes">${data.notes.map((note) => `<div class="muted">・${esc(note)}</div>`).join("")}</div>`;
    }
  </script>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    """ローカルWebサーバーのリクエスト処理。

    このアプリは外部サーバーへデータを送らず、127.0.0.1上でだけ動きます。
    利用者が選択したExcelは /check でこのPythonプロセスへ送られ、
    判定後はメモリ上のJSONとして返します。ファイル保存はしていません。
    """

    def do_GET(self):
        # 画面本体を返します。ルート以外のURLは使いません。
        if self.path not in {"/", "/index.html"}:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        self.respond(HTML_PAGE.encode("utf-8"), "text/html; charset=utf-8")

    def do_POST(self):
        # Excelファイルを受け取り、判定結果JSONを返すAPIです。
        if self.path != "/check":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            # multipart/form-data を標準ライブラリで読み取ります。
            # Python 3.13 では cgi モジュールが削除されたため、email.parser で解析します。
            form = parse_multipart_form(self.headers, self.rfile)
            file_part = form.get("file")
            if not file_part:
                raise ValueError("ファイルが送信されていません。")
            mode = (form.get("mode") or {}).get("text", "auto")
            filename = file_part.get("filename", "")
            data = file_part["data"]

            # Excel -> Course配列 -> シート別判定 -> JSON の順で処理します。
            result = evaluate(read_courses(data, filename), mode)
            self.respond(json.dumps(result, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8")
        except Exception as exc:
            # Excel形式が壊れている、想定外のシート構造などの場合は、
            # ブラウザ側で表示できるエラーメッセージをJSONで返します。
            payload = json.dumps({"error": f"エラー: {exc}"}, ensure_ascii=False).encode("utf-8")
            self.respond(payload, "application/json; charset=utf-8", HTTPStatus.BAD_REQUEST)

    def log_message(self, format, *args):
        # コンソールログが毎回増えると利用者に不要な情報が出るため抑制します。
        return

    def respond(self, body: bytes, content_type: str, status: HTTPStatus = HTTPStatus.OK):
        # HTML/JSONを返す共通処理です。
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def parse_multipart_form(headers, rfile) -> dict:
    """multipart/form-data を cgi 非依存で解析する。

    macOSの新しいPythonでは cgi が削除されているため、HTTP本文に最低限の
    MIMEヘッダーを付けて email.parser に渡します。
    """
    content_type = headers.get("Content-Type", "")
    content_length = int(headers.get("Content-Length", "0"))
    if "multipart/form-data" not in content_type:
        raise ValueError("フォーム形式が正しくありません。")
    body = rfile.read(content_length)
    raw_message = (
        f"Content-Type: {content_type}\r\n"
        "MIME-Version: 1.0\r\n"
        "\r\n"
    ).encode("utf-8") + body
    message = BytesParser(policy=policy.default).parsebytes(raw_message)
    result = {}
    for part in message.iter_parts():
        disposition = part.get("Content-Disposition", "")
        if "form-data" not in disposition:
            continue
        name = part.get_param("name", header="Content-Disposition")
        if not name:
            continue
        data = part.get_payload(decode=True) or b""
        filename = part.get_param("filename", header="Content-Disposition") or ""
        if filename:
            result[name] = {"filename": filename, "data": data}
        else:
            charset = part.get_content_charset() or "utf-8"
            result[name] = {"text": data.decode(charset, errors="replace")}
    return result


def main():
    """アプリ起動処理。

    8765番ポートから順に空きを探してローカルサーバーを起動し、
    少し待ってから既定ブラウザでURLを開きます。
    """
    host = "127.0.0.1"
    port = find_port(host, 8765)
    server = ThreadingHTTPServer((host, port), Handler)
    url = f"http://{host}:{port}/"
    print(f"修了要件チェッカーを起動しました: {url}")
    threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    server.serve_forever()


def find_port(host: str, start: int) -> int:
    """start から20個分のポートを調べ、空いている最初の番号を返す。

    すでに別のアプリや以前起動した本アプリが8765を使っている場合でも、
    8766, 8767... とずらして起動できるようにしています。
    """
    for port in range(start, start + 20):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            if sock.connect_ex((host, port)) != 0:
                return port
    raise RuntimeError("利用可能なポートが見つかりませんでした。")


if __name__ == "__main__":
    main()
